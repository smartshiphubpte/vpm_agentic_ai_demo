"""Inbox drop-folder I/O — pre-voyage / noon Excel (or CSV stand-in).

Supports:
  1) Flat CSV/xlsx demo schema (voyage_number, cp_speed_kn, waypoints, …)
  2) Real SSH Pre-Dep Voyage workbook (sheets: Vessel Details, Voyage Details,
     Waypoints List, CP Terms FWC)

Drop pre-voyage files into VPM_INBOX_DIR/incoming/. Noon → VPM_NOON_INBOX_DIR/incoming/.
"""

from __future__ import annotations

import csv
import io
import re
import shutil
from pathlib import Path
from typing import Any, Literal

from vpm_agents.tools.voyage_registry import normalize_voyage_number

from vpm_agents.tools.folder_layout import FAILED, INCOMING, SENT, ensure_drop_dirs, incoming_dir

MsgKind = Literal["pre_voyage", "noon_report", "unknown"]

_INBOX_SUFFIXES = {".csv", ".xlsx", ".xlsm"}

# Section tags — never treat as field labels (RPM tables use Ballast/Laden as row groups)
_SECTION_TAGS = {
    "ballast",
    "laden",
    "remarks",
    "engine info",
    "general vessel info",
    "charter party info. for voyage",
    "intial voyage info.",
    "initial voyage info.",
    "allowed weather",
    "sample format",
}

_LAT_HEADERS = {"lat", "latitude"}
_LON_HEADERS = {"long", "lon", "lng", "longitude"}
_CONDITIONS = frozenset({"ballast", "laden"})


def _norm_header(h: str) -> str:
    return h.strip().lower().replace(" ", "_")


def _norm_label(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _is_xlsx_name(name: str) -> bool:
    return Path(name).suffix.lower() in {".xlsx", ".xlsm"}


def _is_xlsx(path: Path) -> bool:
    return _is_xlsx_name(path.name)


def _load_wb(source: Path | bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl required for .xlsx — pip install openpyxl") from e
    if isinstance(source, (bytes, bytearray)):
        raw = bytes(source)
        try:
            return load_workbook(io.BytesIO(raw), data_only=True)
        except TypeError:
            return load_workbook(_xlsx_strip_row_merges(raw), data_only=True)
    path = Path(source)
    try:
        return load_workbook(path, data_only=True)
    except TypeError:
        # Excel whole-row merges (`ref="4:4"`) crash openpyxl CellRange.
        return load_workbook(_xlsx_strip_row_merges(path.read_bytes()), data_only=True)


def _xlsx_strip_row_merges(data: bytes):
    """Return a BytesIO xlsx with whole-row merge refs removed."""
    from zipfile import ZIP_DEFLATED, ZipFile

    out = io.BytesIO()
    with ZipFile(io.BytesIO(data), "r") as zin, ZipFile(out, "w", ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            chunk = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                chunk = re.sub(rb'<mergeCell[^>]*ref="\d+:\d+"[^/]*/>', b"", chunk)
            zout.writestr(item, chunk)
    out.seek(0)
    return out


def _row_cap(ws, cap: int) -> int:
    mr = getattr(ws, "max_row", None)
    try:
        mr = int(mr) if mr else 0
    except (TypeError, ValueError):
        mr = 0
    return min(mr, cap) if mr else cap


def _sheet_role(title: str) -> str | None:
    t = _norm_label(title)
    if "waypoint" in t:
        return "waypoints"
    if "voyage" in t:
        return "voyage"
    if "vessel" in t:
        return "vessel"
    if "cp term" in t or "fwc" in t or "charter party" in t:
        return "cp"
    return None


def _sheet_map(wb) -> dict[str, Any]:
    """Map logical roles → worksheet by title keywords, not a fixed ship template."""
    out: dict[str, Any] = {}
    for ws in wb.worksheets:
        role = _sheet_role(ws.title)
        if role and role not in out:
            out[role] = ws
    return out


def is_predep_workbook(path: Path | None = None, *, data: bytes | None = None, name: str = "") -> bool:
    label = name or (path.name if path else "")
    if data is not None:
        if not _is_xlsx_name(label):
            return False
        wb = _load_wb(data)
    else:
        if path is None or not _is_xlsx(path):
            return False
        wb = _load_wb(path)
    try:
        sheets = _sheet_map(wb)
        if "waypoints" in sheets:
            return True
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            return False
        # Combined noon Excel: Voyage_Number + Latitude columns — not a Pre-Dep form
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
        cols = {_norm_header(str(h)) for h in header if h is not None}
        if _kind_from_cols(cols) == "noon_report":
            return False
        return _find_lat_lon_cols(ws)[0] is not None
    finally:
        wb.close()


def _as_float(val: Any) -> float | None:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
        return float(s)
    return None


def _sample_format_col(ws) -> int | None:
    """Column index of the 'Sample format' header (template values, never ingest)."""
    for row in ws.iter_rows(min_row=1, max_row=_row_cap(ws, 20), max_col=12, values_only=True):
        for i, c in enumerate(row):
            if c is not None and _norm_label(c) == "sample format":
                return i
    return None


def _is_field_label(cell: Any) -> bool:
    if cell is None or not isinstance(cell, str):
        return False
    key = _norm_label(cell)
    if len(key) < 3 or key in _SECTION_TAGS:
        return False
    if key in _LAT_HEADERS or key in _LON_HEADERS:
        return False
    return bool(re.search(r"[a-zA-Z]", key))


def _sheet_fields(ws) -> dict[str, list[Any]]:
    """Label → values to the right. Works for col-C labels or any leading field name.

    The Pre-Dep template puts real values next to the label and a 'Sample format'
    column further right; that column is ignored.
    """
    found: dict[str, list[Any]] = {}
    sample_col = _sample_format_col(ws)
    max_row = _row_cap(ws, 80)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=12, values_only=True):
        cells = list(row)
        label_i = None
        for i, c in enumerate(cells[:3]):
            if _is_field_label(c):
                label_i = i
        if label_i is None:
            continue
        rest: list[Any] = []
        for i, c in enumerate(cells):
            if i <= label_i:
                continue
            if sample_col is not None and i == sample_col:
                continue
            if c is not None and str(c).strip() != "":
                rest.append(c)
        if not rest:
            continue
        found.setdefault(_norm_label(cells[label_i]), rest)
    return found


def _merge_fields(wb, sheets: dict[str, Any]) -> dict[str, list[Any]]:
    fields: dict[str, list[Any]] = {}
    for role in ("voyage", "cp", "vessel"):
        ws = sheets.get(role)
        if ws is None:
            continue
        for k, v in _sheet_fields(ws).items():
            fields.setdefault(k, v)
    for ws in wb.worksheets:
        for k, v in _sheet_fields(ws).items():
            fields.setdefault(k, v)
    return fields


_SKIP_FIELD_EXTRA = ("allowance", "variance", "allowed", "tolerance", "tolerence")


def _best_field(
    mapping: dict[str, list[Any]],
    *needles: str,
    skip_extra: tuple[str, ...] = _SKIP_FIELD_EXTRA,
) -> tuple[str, list[Any]] | None:
    """Best label containing a needle; skip allowance/tolerance extras unless skip_extra=()."""
    ranked: list[tuple[int, int, str, list[Any]]] = []
    for needle in needles:
        n = _norm_label(needle)
        for k, vals in mapping.items():
            if n not in k:
                continue
            extra = k.replace(n, "", 1).strip()
            if extra and any(w in extra for w in skip_extra):
                continue
            ranked.append((len(extra), -len(n), k, vals))
    if not ranked:
        return None
    ranked.sort()
    return ranked[0][2], ranked[0][3]


def _first_number(vals: list[Any]) -> float | None:
    for v in vals:
        n = _as_float(v)
        if n is not None:
            return n
    return None


def _fmt_cells(vals: list[Any] | None) -> str:
    if not vals:
        return "(empty)"
    return ", ".join(repr(v) for v in vals[:8])


def _pick_enum(
    mapping: dict[str, list[Any]],
    *needles: str,
    allowed: frozenset[str],
) -> str | None:
    """First cell value whose text contains an allowed token (Ballast/Laden, etc.)."""
    hit = _best_field(mapping, *needles)
    if not hit:
        return None
    for v in hit[1]:
        if v is None:
            continue
        s = _norm_label(str(v))
        for a in allowed:
            if a in s:
                return a.capitalize()
    return None


def _pick(mapping: dict[str, list[Any]], *needles: str, numeric: bool = False) -> Any:
    """Best label containing a needle; skip allowance/variance extras unless asked."""
    hit = _best_field(mapping, *needles)
    if not hit:
        return None
    _label, vals = hit
    if numeric:
        return _first_number(vals)
    return vals[0]


def _numeric_or_issue(
    mapping: dict[str, list[Any]],
    title: str,
    *needles: str,
    required: bool = True,
    skip_extra: tuple[str, ...] = _SKIP_FIELD_EXTRA,
) -> tuple[float | None, str | None]:
    """Parse a numeric field. On failure return a reject line naming the Excel label + cells."""
    hit = _best_field(mapping, *needles, skip_extra=skip_extra)
    looked = ", ".join(repr(n) for n in needles)
    if hit is None:
        if not required:
            return None, None
        return None, f"{title}: field not found (looked for labels containing {looked})"
    label, vals = hit
    num = _first_number(vals)
    if num is not None:
        return num, None
    why = (
        f"{title}: rejected — Excel label {label!r} has no numeric value "
        f"(got {_fmt_cells(vals)}). Need a number; Ballast/Laden/text is not enough."
    )
    if not required:
        return None, why
    return None, why


def _find_lat_lon_cols(ws) -> tuple[int | None, int | None, int, int]:
    """Return (lat_col, lon_col, name_col, first_data_row)."""
    if ws is None:
        return None, None, 0, 1
    lat_col = lon_col = None
    name_col = 0
    header_row = 1
    for r_i, row in enumerate(ws.iter_rows(max_row=_row_cap(ws, 25), max_col=12, values_only=True), 1):
        for ci, c in enumerate(row):
            if c is None:
                continue
            n = _norm_label(c).rstrip(".")
            if n in _LAT_HEADERS:
                lat_col = ci
                header_row = r_i
            elif n in _LON_HEADERS:
                lon_col = ci
                header_row = r_i
            elif n in {"wpt no", "wpt", "waypoint", "name"}:
                name_col = ci
        if lat_col is not None and lon_col is not None:
            return lat_col, lon_col, name_col, header_row + 1
    return None, None, 0, 1


def _parse_waypoints(ws) -> tuple[list[list[float]], list[str]]:
    lat_col, lon_col, name_col, start = _find_lat_lon_cols(ws)
    if lat_col is None or lon_col is None:
        return [], []
    waypoints: list[list[float]] = []
    wp_names: list[str] = []
    originals: list[tuple[Any, Any]] = []
    bad: list[str] = []
    need = max(lat_col, lon_col, name_col) + 1
    iter_kw: dict[str, Any] = {"min_row": start, "max_col": need, "values_only": True}
    mr = getattr(ws, "max_row", None)
    try:
        if mr:
            iter_kw["max_row"] = int(mr)
    except (TypeError, ValueError):
        pass
    for row_i, row in enumerate(ws.iter_rows(**iter_kw), start):
        cells = list(row) + [None] * need
        lat_s, lon_s = cells[lat_col], cells[lon_col]
        if lat_s is None or lon_s is None:
            continue
        if str(lat_s).strip() == "" or str(lon_s).strip() == "":
            continue
        if _norm_label(str(lat_s)) in _LAT_HEADERS:
            continue
        try:
            lat = parse_dm_coordinate(lat_s)
            lon = parse_dm_coordinate(lon_s)
        except ValueError as e:
            bad.append(f"row {row_i}: {e}")
            continue
        originals.append((lat_s, lon_s))
        waypoints.append([lat, lon])
        name = cells[name_col]
        wp_names.append("" if name is None else str(name).strip())
    if bad:
        raise ValueError("unparseable master-route coordinates:\n- " + "\n- ".join(bad))
    assert_converted_match_master(originals, waypoints)
    return waypoints, wp_names


_COORD_EPS = 1e-5  # ~1.1 m — conversion must not drift off the master point


def parse_dm_coordinate(text: Any) -> float:
    """Degrees (DMS / DM / hemisphere) or decimal → signed decimal degrees."""
    if isinstance(text, bool) or text is None:
        raise ValueError(f"bad coordinate: {text!r}")
    if isinstance(text, (int, float)):
        return round(float(text), 6)
    s = str(text).strip().upper().replace("''", '"').replace("′", "'").replace("″", '"').replace("º", "°")
    s = re.sub(r"\s+", " ", s)
    m = re.match(
        r"^([+-]?\d+(?:\.\d+)?)[°\s-]+(\d+(?:\.\d+)?)['\s-]+(\d+(?:\.\d+)?)[\"'\s]*([NSEW])$",
        s,
    )
    if m:
        val = float(m.group(1)) + float(m.group(2)) / 60.0 + float(m.group(3)) / 3600.0
        if m.group(4) in ("S", "W"):
            val = -val
        return round(val, 6)
    m = re.match(
        r"^([+-]?\d+(?:\.\d+)?)[°\s-]+(\d+(?:\.\d+)?)['\"]?\s*([NSEW])$",
        s,
    )
    if m:
        val = float(m.group(1)) + float(m.group(2)) / 60.0
        if m.group(3) in ("S", "W"):
            val = -val
        return round(val, 6)
    m = re.match(r"^([+-]?\d+(?:\.\d+)?)\s*°?\s*([NSEW])$", s)
    if m:
        val = float(m.group(1))
        if m.group(2) in ("S", "W"):
            val = -val
        return round(val, 6)
    try:
        return round(float(s.replace("°", "").strip()), 6)
    except ValueError as e:
        raise ValueError(f"bad coordinate: {text!r}") from e


def _format_dm(val: float, *, lon: bool) -> str:
    """Canonical deg + decimal-minutes for round-trip checks."""
    hemi = ("E" if val >= 0 else "W") if lon else ("N" if val >= 0 else "S")
    v = abs(float(val))
    deg = int(v)
    minutes = (v - deg) * 60.0
    if minutes >= 59.99995:
        deg += 1
        minutes = 0.0
    return f"{deg} {minutes:.4f} {hemi}"


def assert_converted_match_master(
    originals: list[tuple[Any, Any]],
    decimals: list[list[float]],
) -> None:
    """Converted decimals must be 1:1 with master degree points (same place, round-trip)."""
    if len(originals) != len(decimals):
        raise ValueError(
            f"converted waypoint count {len(decimals)} != master count {len(originals)}"
        )
    for i, ((raw_lat, raw_lon), pair) in enumerate(zip(originals, decimals), 1):
        if not pair or len(pair) < 2:
            raise ValueError(f"waypoint {i}: missing converted lat/lon")
        lat, lon = float(pair[0]), float(pair[1])
        if abs(lat) > 90.0 or abs(lon) > 180.0:
            raise ValueError(f"waypoint {i}: decimal out of range lat={lat} lon={lon}")
        master_lat = parse_dm_coordinate(raw_lat)
        master_lon = parse_dm_coordinate(raw_lon)
        if abs(lat - master_lat) > _COORD_EPS or abs(lon - master_lon) > _COORD_EPS:
            raise ValueError(
                f"waypoint {i}: converted ({lat}, {lon}) does not match master "
                f"({master_lat}, {master_lon}) from {raw_lat!r}, {raw_lon!r}"
            )
        back_lat = parse_dm_coordinate(_format_dm(lat, lon=False))
        back_lon = parse_dm_coordinate(_format_dm(lon, lon=True))
        if abs(back_lat - master_lat) > _COORD_EPS or abs(back_lon - master_lon) > _COORD_EPS:
            raise ValueError(
                f"waypoint {i}: degree↔decimal round-trip drifted from master "
                f"({master_lat}, {master_lon})"
            )


def parse_predep_workbook(
    path: Path | None = None, *, data: bytes | None = None, name: str = ""
) -> dict[str, Any]:
    """Parse a Pre-Dep-style workbook: any sheet titles, label/value rows, lat/long table."""
    label = name or (path.name if path else "workbook")
    wb = _load_wb(data if data is not None else path)  # type: ignore[arg-type]
    try:
        sheets = _sheet_map(wb)
        fields = _merge_fields(wb, sheets)
        wp_ws = sheets.get("waypoints") or next(
            (ws for ws in wb.worksheets if _find_lat_lon_cols(ws)[0] is not None),
            None,
        )
        missing: list[str] = []
        if wp_ws is None:
            missing.append(
                "Waypoints List sheet with Lat and Long (or Latitude/Longitude) header columns"
            )

        voyage_number = _pick(fields, "voyage number", "voyage no")
        source_port = _pick(fields, "departure port", "from port", "origin")
        dest_port = _pick(fields, "destination port", "dest port", "arrival port")
        vessel_name = _pick(fields, "vessel name", "ship name")
        imo = _pick(fields, "imo no", "imo number")
        # CP speed/cons come from consumption_speed_data/, not the Pre-Dep CP sheet.

        vn_hit = _best_field(fields, "voyage number", "voyage no")
        if voyage_number is None:
            missing.append(
                "Voyage Number: field not found or empty"
                + (f" (label {vn_hit[0]!r} cells {_fmt_cells(vn_hit[1])})" if vn_hit else "")
            )
        sp_hit = _best_field(fields, "departure port", "from port", "origin")
        if source_port is None:
            missing.append(
                "Departure port: field not found or empty"
                + (f" (label {sp_hit[0]!r} cells {_fmt_cells(sp_hit[1])})" if sp_hit else "")
            )
        dp_hit = _best_field(fields, "destination port", "dest port", "arrival port")
        if dest_port is None:
            missing.append(
                "Destination port: field not found or empty"
                + (f" (label {dp_hit[0]!r} cells {_fmt_cells(dp_hit[1])})" if dp_hit else "")
            )
        if not (vessel_name or imo):
            missing.append("Vessel Name or IMO number: both empty (must match client shipping_db.ship)")
        cp_speed = cp_consumption = None
        try:
            from inbox_agent.vessel_matrix import require_speed_cons

            cp_speed, cp_consumption = require_speed_cons(
                "" if vessel_name is None else str(vessel_name)
            )
        except ValueError as e:
            missing.append(str(e))

        waypoints, wp_names = _parse_waypoints(wp_ws) if wp_ws is not None else ([], [])
        if wp_ws is not None and len(waypoints) < 2:
            missing.append("at least 2 waypoint rows with parseable Lat/Long values")
        if missing:
            raise ValueError(f"{label}: not a valid pre-voyage report.\n" + "\n".join(missing))

        vessel_id = str(imo).strip() if imo is not None else (str(vessel_name).strip() if vessel_name else "")
        cond = _pick_enum(fields, "condition (ballast/laden)", "condition", allowed=_CONDITIONS)
        etd = _pick(fields, "estimated departure time", "estimated date of departure", "etd")
        eta = _pick(fields, "estimated arrival time", "estimated date of arrival", "eta")
        record = {
            "voyage_number": normalize_voyage_number(str(voyage_number)),
            "vessel_id": vessel_id,
            "vessel_name": "" if vessel_name is None else str(vessel_name).strip(),
            "source_port": str(source_port).strip(),
            "dest_port": str(dest_port).strip(),
            "cp_speed_kn": float(cp_speed),
            "cp_consumption_mt_day": cp_consumption,
            "alert_emails": [],
            "master_waypoints": waypoints,
            "waypoint_names": wp_names,
            "source_file": label,
            "format": "predep_xlsx",
            "condition": cond or "",
            "etd": "" if etd is None else str(etd).strip(),
            "eta": "" if eta is None else str(eta).strip(),
            "displacement": _pick(fields, "displacement", numeric=True),
            "cargo_weight": _pick(fields, "cargo weight", numeric=True),
            "max_draft_on_departure": _pick(
                fields, "max draft upon departure", "max draft on departure", numeric=True
            ),
            "voyage_priority": str(_pick(fields, "voyage priority") or "").strip(),
        }
        from inbox_agent.validate import validate_pre_voyage

        extra = validate_pre_voyage(record)
        if extra:
            raise ValueError(f"{label}: not a valid pre-voyage report.\n" + "\n".join(extra))
        return record
    finally:
        wb.close()


def _read_table(
    path: Path | None = None, *, data: bytes | None = None, filename: str = ""
) -> tuple[list[str], list[str]]:
    """Return (headers, first_data_row_values) from flat csv or single-sheet xlsx."""
    label = filename or (path.name if path else "file")
    suffix = Path(label).suffix.lower()
    if path is not None:
        suffix = path.suffix.lower()
        label = path.name
    if suffix == ".csv":
        text = (
            data.decode("utf-8-sig")
            if data is not None
            else Path(path).read_text(encoding="utf-8-sig")  # type: ignore[arg-type]
        )
        rows = list(csv.reader(io.StringIO(text)))
        if len(rows) < 2:
            raise ValueError(f"{label}: need header + one data row")
        return [_norm_header(h) for h in rows[0]], [c.strip() for c in rows[1]]

    if suffix in {".xlsx", ".xlsm"}:
        wb = _load_wb(data if data is not None else path)  # type: ignore[arg-type]
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            data_row = next(rows_iter, None)
            if not header_row or not data_row:
                raise ValueError(f"{label}: need header + one data row")
            headers = [_norm_header(str(h)) for h in header_row if h is not None]
            values = ["" if v is None else str(v).strip() for v in data_row[: len(headers)]]
            return headers, values
        finally:
            wb.close()

    raise ValueError(f"unsupported inbox file type: {suffix}")


def _kind_from_cols(cols: set[str]) -> MsgKind | None:
    if "waypoints" in cols and "cp_speed_kn" in cols:
        return "pre_voyage"
    has_lat = bool(cols & {"lat", "latitude"})
    has_lon = bool(cols & {"lon", "long", "lng", "longitude"})
    if has_lat and has_lon and "voyage_number" in cols:
        return "noon_report"
    return None


def classify_inbox_file(
    path: str | Path | None = None, *, data: bytes | None = None, filename: str = ""
) -> MsgKind:
    label = filename or (Path(path).name if path else "")
    name_l = label.lower()
    xlsx = _is_xlsx_name(label) if (data is not None or label) else False
    if path and not data:
        xlsx = _is_xlsx(Path(path))

    # Multi-sheet Pre-Dep first (waypoint sheet). Noon workbooks also have Lat/Long
    # but usually no "waypoints" sheet — those must classify as noon_report.
    if xlsx:
        try:
            if is_predep_workbook(Path(path) if path else None, data=data, name=label):
                return "pre_voyage"
        except Exception:
            pass

    try:
        headers, _ = _read_table(Path(path) if path else None, data=data, filename=label)
    except Exception:
        headers = []
    kind = _kind_from_cols(set(headers))
    if kind:
        return kind

    if xlsx and (
        name_l.startswith("pre-dep") or "pre-dep voyage" in name_l or "pre_dep" in name_l
    ):
        return "pre_voyage"  # let parse raise a clear error
    return "unknown"


def _parse_waypoint_field(wp_raw: str) -> list[list[float]]:
    originals: list[tuple[Any, Any]] = []
    waypoints: list[list[float]] = []
    for pair in wp_raw.split(";"):
        pair = pair.strip()
        if not pair:
            continue
        lat_s, lon_s = pair.split(",", 1)
        originals.append((lat_s.strip(), lon_s.strip()))
        waypoints.append([parse_dm_coordinate(lat_s), parse_dm_coordinate(lon_s)])
    if len(waypoints) < 2:
        raise ValueError("pre-voyage needs ≥2 waypoints")
    assert_converted_match_master(originals, waypoints)
    return waypoints


def parse_pre_voyage(
    path: str | Path | None = None, *, data: bytes | None = None, filename: str = ""
) -> dict[str, Any]:
    """Parse flat CSV/xlsx demo schema OR real Pre-Dep Voyage workbook."""
    label = filename or (Path(path).name if path else "attachment")
    if data is not None:
        if _is_xlsx_name(label) and is_predep_workbook(data=data, name=label):
            return parse_predep_workbook(data=data, name=label)
        if Path(label).suffix.lower() == ".csv":
            text = data.decode("utf-8-sig").strip()
        else:
            headers, values = _read_table(data=data, filename=label)
            return _parse_flat_row(dict(zip(headers, values)), label)
    else:
        path = Path(path)  # type: ignore[arg-type]
        label = path.name
        if _is_xlsx(path) and is_predep_workbook(path):
            return parse_predep_workbook(path)
        if path.suffix.lower() == ".csv":
            text = path.read_text(encoding="utf-8-sig").strip()
        else:
            headers, values = _read_table(path)
            return _parse_flat_row(dict(zip(headers, values)), str(path))

    lines = text.splitlines()
    if len(lines) < 2:
        raise ValueError(f"{label}: need header + one data row")
    rows = list(csv.reader(io.StringIO("\n".join(lines))))
    if len(rows) < 2:
        raise ValueError(f"{label}: need header + one data row")
    headers = [_norm_header(h) for h in rows[0]]
    values = [("" if v is None else str(v).strip()) for v in rows[1][: len(headers)]]
    row = dict(zip(headers, values))
    src = label if data is not None else str(path)
    return _parse_flat_row(row, src)


def _parse_flat_row(row: dict[str, Any], source: str) -> dict[str, Any]:
    missing = [
        key
        for key in ("voyage_number", "vessel_id", "source_port", "dest_port", "cp_speed_kn", "waypoints")
        if key not in row or not row[key]
    ]
    if missing:
        raise ValueError(
            f"{Path(source).name}: not a valid pre-voyage report. Missing columns/values:\n- "
            + "\n- ".join(missing)
        )
    emails = [e for e in str(row.get("alert_emails", "") or "").split("|") if e.strip()]
    cons_raw = row.get("cp_consumption_mt_day") or row.get("cp_consumption")
    cons = None
    if cons_raw not in (None, ""):
        try:
            cons = float(cons_raw)
        except ValueError:
            cons = None
    return {
        "voyage_number": normalize_voyage_number(row["voyage_number"]),
        "vessel_id": row["vessel_id"],
        "vessel_name": row.get("vessel_name", ""),
        "source_port": row["source_port"],
        "dest_port": row["dest_port"],
        "cp_speed_kn": float(row["cp_speed_kn"]),
        "cp_consumption_mt_day": cons,
        "alert_emails": emails,
        "master_waypoints": _parse_waypoint_field(row["waypoints"]),
        "waypoint_names": [
            n.strip() for n in str(row.get("waypoint_names") or "").split("|") if n.strip()
        ],
        "source_file": source,
        "format": "flat",
        "condition": str(row.get("condition") or "").strip(),
        "displacement": row.get("displacement"),
        "cargo_weight": row.get("cargo_weight"),
        "max_draft_on_departure": row.get("max_draft_on_departure"),
        "etd": str(row.get("etd") or row.get("estimated_departure_time") or row.get("estimated_date_of_departure") or "").strip(),
        "eta": str(row.get("eta") or row.get("estimated_arrival_time") or row.get("estimated_date_of_arrival") or "").strip(),
    }


def try_parse_pre_voyage(*, filename: str, data: bytes) -> tuple[dict[str, Any] | None, list[str]]:
    """Parse attachment bytes. Returns (record, []) or (None, human-readable reject reasons)."""
    name = filename or "attachment"
    suf = Path(name).suffix.lower()
    if suf not in _INBOX_SUFFIXES:
        return None, [
            f"{name}: unsupported type {suf or '(no extension)'}. "
            "Attach a Pre-Dep .xlsx/.xlsm or a pre-voyage .csv."
        ]
    kind = classify_inbox_file(data=data, filename=name)
    if kind == "noon_report":
        return None, [
            f"{name}: classified as a noon-position report (Latitude/Longitude + Voyage_Number), "
            "not a pre-voyage / Pre-Dep workbook."
        ]
    try:
        rec = parse_pre_voyage(data=data, filename=name)
    except Exception as e:
        lines = [ln.strip() for ln in str(e).splitlines() if ln.strip()]
        if not lines:
            return None, [f"{name}: parse failed"]
        return None, [ln if ln.lower().startswith(name.lower()) else f"{name}: {ln}" for ln in lines]
    from inbox_agent.validate import validate_pre_voyage, vessel_register_issues

    field_issues = validate_pre_voyage(rec)
    if field_issues:
        return None, [f"{name}: {line}" for line in field_issues]
    reg = vessel_register_issues(rec)
    if reg:
        return None, [f"{name}: {line}" for line in reg]
    return rec, []


def parse_noon_report(path: str | Path) -> dict[str, Any]:
    """Columns: voyage_number, lat, lon  (optional observed_at)."""
    headers, values = _read_table(Path(path))
    row = dict(zip(headers, values))
    for key in ("voyage_number", "lat", "lon"):
        if key not in row or not row[key]:
            raise ValueError(f"noon report missing {key}")
    return {
        "voyage_number": normalize_voyage_number(row["voyage_number"]),
        "lat": float(row["lat"]),
        "lon": float(row["lon"]),
        "observed_at": row.get("observed_at") or None,
        "report_type": str(row.get("report_type") or "").strip() or None,
        "source_file": str(path),
    }


def list_inbox(inbox_dir: str | Path) -> list[Path]:
    inbox = Path(inbox_dir)
    ensure_drop_dirs(inbox)
    drop = incoming_dir(inbox)
    return sorted(
        p
        for p in drop.iterdir()
        if p.is_file() and p.suffix.lower() in _INBOX_SUFFIXES and not p.name.startswith(".")
    )


def relocate_inbox_file(path: Path, dest_dir: str | Path) -> Path:
    path = Path(path)
    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / path.name
    if dest.exists():
        dest = dest_dir / f"{path.stem}_{path.stat().st_mtime_ns}{path.suffix}"
    shutil.move(str(path), str(dest))
    return dest


def archive_inbox_file(path: Path, dest_subdir: str = SENT) -> Path:
    return relocate_inbox_file(path, path.parent.parent / dest_subdir)


if __name__ == "__main__":
    assert parse_dm_coordinate("12 39.812 N") == round(12 + 39.812 / 60.0, 6)
    assert parse_dm_coordinate("109°24.414'E") == round(109 + 24.414 / 60.0, 6)
    assert parse_dm_coordinate("35°13'6''S") == round(-(35 + 13 / 60.0 + 6 / 3600.0), 6)
    assert parse_dm_coordinate("1.25") == 1.25
    lat_dm, lon_dm = "12 39.812 N", "109 24.414 E"
    dec = [[parse_dm_coordinate(lat_dm), parse_dm_coordinate(lon_dm)]]
    assert_converted_match_master([(lat_dm, lon_dm)], dec)
    try:
        assert_converted_match_master([(lat_dm, lon_dm)], [[12.0, 109.0]])
        raise AssertionError("mismatch should fail")
    except ValueError:
        pass
    assert _as_float("Ballast") is None
    assert _as_float(10.7) == 10.7
    assert _as_float("12.5") == 12.5
    mapping = {
        "cp speed in kts/day": ["Ballast", 10.7, "Laden", 12.5],
        "cp speed allowance variance (kts)": [0.5],
        "cp consumptions(total) in mts/day": ["Ballast", 25],
    }
    assert _pick(mapping, "cp speed", numeric=True) == 10.7
    assert _pick(mapping, "cp consumption", numeric=True) == 25
    spd, spd_err = _numeric_or_issue(mapping, "CP Speed (knots)", "cp speed")
    assert spd == 10.7 and spd_err is None
    _, lol_err = _numeric_or_issue({"cp speed in kts/day": ["LOL"]}, "CP Speed (knots)", "cp speed")
    assert lol_err and "LOL" in lol_err

    class _SampleWs:
        max_row = 8

        def iter_rows(self, min_row=1, max_row=20, max_col=12, values_only=True):
            yield (None, "Charter Party Info. For Voyage", None, None, None, None, "Sample format")
            yield (None, None, "CP Speed in kts/day", "LOL", None, None, 12.5)

    assert _sample_format_col(_SampleWs()) == 6
    assert _sheet_fields(_SampleWs())["cp speed in kts/day"] == ["LOL"]
    assert _kind_from_cols({"voyage_number", "latitude", "longitude"}) == "noon_report"
    assert _kind_from_cols({"voyage_number", "lat", "lon"}) == "noon_report"
    assert _kind_from_cols({"waypoints", "cp_speed_kn"}) == "pre_voyage"
    stripped = re.sub(rb'<mergeCell[^>]*ref="\d+:\d+"[^/]*/>', b"", b'<mergeCell ref="4:4"/><mergeCell ref="A1:B1"/>')
    assert b'ref="4:4"' not in stripped and b'ref="A1:B1"' in stripped
    sample = Path(__file__).resolve().parents[1] / "samples" / "inbox" / "pre_voyage.csv"
    blob = sample.read_bytes()
    rec, errs = try_parse_pre_voyage(filename=sample.name, data=blob)
    assert not errs and rec and rec["voyage_number"]
    bad, bad_errs = try_parse_pre_voyage(filename="note.txt", data=b"hello")
    assert bad is None and bad_errs
    print("inbox_io self-check ok")
