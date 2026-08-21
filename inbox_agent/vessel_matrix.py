"""Vessel name → CP speed / daily fuel from inbox_agent/consumption_speed_data/."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

_DIR = Path(__file__).resolve().parent / "consumption_speed_data"
_cache: tuple[float, dict[str, tuple[float, float]]] | None = None


def _norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _as_float(val: Any) -> float | None:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
        return float(s)
    return None


def _load() -> dict[str, tuple[float, float]]:
    global _cache
    files = sorted(_DIR.glob("*.xlsx"))
    mtime = max((f.stat().st_mtime for f in files), default=0.0)
    if _cache is not None and _cache[0] == mtime:
        return _cache[1]
    table: dict[str, tuple[float, float]] = {}
    if not files:
        _cache = (mtime, table)
        return table
    from openpyxl import load_workbook

    for path in files:
        wb = load_workbook(path, data_only=True)
        try:
            for ws in wb.worksheets:
                header_i: dict[str, int] = {}
                for row in ws.iter_rows(values_only=True):
                    cells = list(row)
                    if not header_i:
                        labels = [_norm_name(c) if c is not None else "" for c in cells]
                        if "vessel name" not in labels or "speed" not in labels:
                            continue
                        header_i = {lab: i for i, lab in enumerate(labels) if lab}
                        continue
                    ni = header_i.get("vessel name")
                    if ni is None or ni >= len(cells) or cells[ni] is None:
                        continue
                    name = _norm_name(str(cells[ni]))
                    if not name or name.startswith("note"):
                        continue
                    speed = _as_float(cells[header_i["speed"]]) if "speed" in header_i else None
                    total_i = header_i.get("total")
                    cons = _as_float(cells[total_i]) if total_i is not None and total_i < len(cells) else None
                    if cons is None:
                        hsfo_i = next((header_i[k] for k in header_i if "hsfo" in k), None)
                        mgo_i = next((header_i[k] for k in header_i if "mgo" in k), None)
                        hsfo = _as_float(cells[hsfo_i]) if hsfo_i is not None and hsfo_i < len(cells) else None
                        mgo = _as_float(cells[mgo_i]) if mgo_i is not None and mgo_i < len(cells) else None
                        if hsfo is not None and mgo is not None:
                            cons = hsfo + mgo
                        else:
                            cons = hsfo
                    if speed is None or cons is None:
                        continue
                    table[name] = (speed, cons)
        finally:
            wb.close()
    _cache = (mtime, table)
    return table


def lookup_speed_cons(vessel_name: str) -> tuple[float, float] | None:
    """Return (speed_kn, fuel_mt_day) or None if the vessel is missing / NA."""
    key = _norm_name(vessel_name)
    if not key:
        return None
    return _load().get(key)


def require_speed_cons(vessel_name: str) -> tuple[float, float]:
    """Lookup or raise ValueError naming the vessel (for Pre-Dep reject mail)."""
    name = str(vessel_name or "").strip()
    if not name:
        raise ValueError(
            "Vessel Name: empty — needed to look up speed and fuel consumption "
            f"in {_DIR.name}/"
        )
    hit = lookup_speed_cons(name)
    if hit is None:
        raise ValueError(
            f"Vessel {name!r}: not in consumption/speed matrix "
            f"({_DIR.name}/) or speed/fuel is NA"
        )
    return hit


if __name__ == "__main__":
    assert lookup_speed_cons("Batavia Express") == (11.0, 21.43)
    assert lookup_speed_cons("brazil express") == (11.75, 22.14)
    assert lookup_speed_cons("ASIA UNITY") is None
    assert lookup_speed_cons("no such ship") is None
    try:
        require_speed_cons("ASIA UNITY")
        raise AssertionError("NA vessel should reject")
    except ValueError as e:
        assert "ASIA UNITY" in str(e)
    print("vessel_matrix self-check ok")
