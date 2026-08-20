"""Parse combined noon-report Excel workbooks (multi-row, BE-style columns)."""

from __future__ import annotations

import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from vpm_agents.tools.voyage_registry import normalize_voyage_number


def parse_dms_coordinate(text: str) -> float:
    """Parse 35°13'6''N, 12 39.812 N, or decimal."""
    from inbox_agent.parse import parse_dm_coordinate

    return parse_dm_coordinate(text)


def _norm_col(h: str) -> str:
    return re.sub(r"\s+", "_", h.strip().lower())


def noon_observed_sort_key(row: dict[str, Any]) -> tuple[int, float, str]:
    """Sort key for oldest-first noon processing."""
    raw = row.get("observed_at")
    if raw:
        try:
            dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return (0, dt.timestamp(), str(row.get("noon_id") or ""))
        except ValueError:
            pass
    return (1, 0.0, str(row.get("noon_id") or ""))


def sort_noon_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return rows oldest observed_at first (stable tie-break on noon_id)."""
    return sorted(rows, key=noon_observed_sort_key)


def noon_row_id(row: dict[str, Any]) -> str:
    rid = row.get("report_id") or row.get("id")
    if rid is not None and str(rid).strip():
        return f"report:{rid}"
    blob = f"{row.get('voyage_number')}|{row.get('observed_at')}|{row.get('lat')}|{row.get('lon')}"
    return "hash:" + hashlib.sha256(blob.encode()).hexdigest()[:16]


def parse_noon_excel(path: str | Path) -> list[dict[str, Any]]:
    """Read all noon rows from a combined Excel workbook."""
    path = Path(path)
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl required — pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        headers = [_norm_col(str(h)) if h is not None else "" for h in header_row]
        col = {h: i for i, h in enumerate(headers) if h}

        def cell(row: tuple, *names: str) -> Any:
            for name in names:
                idx = col.get(name)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    if v is not None and str(v).strip() != "":
                        return v
            return None

        out: list[dict[str, Any]] = []
        for row in rows_iter:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            lat_raw = cell(row, "latitude", "lat")
            lon_raw = cell(row, "longitude", "lon")
            voy = cell(row, "voyage_number")
            if lat_raw is None or lon_raw is None or voy is None:
                continue
            try:
                lat = parse_dms_coordinate(lat_raw)
                lon = parse_dms_coordinate(lon_raw)
            except ValueError:
                continue

            observed = cell(row, "report_date_time", "observed_at", "report_date_time_local")
            report_type = str(cell(row, "report_type") or "").strip()
            nd = _noonreportdata_from_row(lambda *names: cell(row, *names))
            record = {
                "voyage_number": normalize_voyage_number(str(voy)),
                "vessel_name": str(cell(row, "vessel_name") or "").strip(),
                "lat": lat,
                "lon": lon,
                "observed_at": str(observed).strip() if observed else None,
                "report_type": report_type,
                "avg_speed_kn": _float_or_none(
                    cell(row, "avg_speed", "average_speed_since_sov", "log_speed")
                ),
                "report_id": cell(row, "report_id", "id"),
                "source_file": str(path),
                "source": "excel",
                # BE-shaped row for EOV formula engine (local or live)
                "eov_row": {
                    "reporttype": report_type,
                    "utcTime": str(observed).strip() if observed else None,
                    "noonreportdata": nd,
                    "lat": lat,
                    "lon": lon,
                },
            }
            record["noon_id"] = noon_row_id(record)
            out.append(record)
        out.reverse()  # sheet is newest-first; process oldest noon first
        return out
    finally:
        wb.close()


def _float_or_none(v: Any) -> float | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# Columns the EOV formula engine / tables need (subset of BE noon workbook).
_EOV_ND_KEYS = (
    "ME_Running_Hrs",
    "ME_RPM",
    "Slip",
    "Distance",
    "Distance_Covered_Since_SOV",
    "Total_ME_Running_Hrs_Since_SOV",
    "Avg_Speed",
    "Wind_Speed",
    "Wind_Force",
    "Beaufort_Scale",
    "Sea_Height",
    "Swell_Height",
    "Wave_Height",
    "Current_Velocity",
    "Current_Direction",
    "Total_HFOME_Consumed_In_MT",
    "Total_VLSFOME_Consumed_In_MT",
    "Total_HFOAE_Consumed_In_MT",
    "Total_VLSFOAX_Consumed_In_MT",
    "Total_HFOBLR_Consumed_In_MT",
    "Total_VLSFOBLR_Consumed_In_MT",
    "Total_HFO_Consumed_In_MT",
    "Total_VLSFO_Consumed_In_MT",
    "Total_LSMGO_Consumed_In_MT",
    "Total_ULSGO_Consumed_In_MT",
    "Total_VLSGO_Consumed_In_MT",
    "Total_LSMGOME_Consumed_In_MT",
    "Total_LSMGOAE_Consumed_In_MT",
    "Total_LSMGOBLR_Consumed_In_MT",
    "Total_ULSGOME_Consumed_In_MT",
    "Total_ULSGOAE_Consumed_In_MT",
    "Total_ULSGOBLR_Consumed_In_MT",
    "Total_VLSGOME_Consumed_In_MT",
    "Total_VLSGOAX_Consumed_In_MT",
    "Total_VLSGOBLR_Consumed_In_MT",
    "Remaining_On_Board_HFO_In_MT",
    "Remaining_On_Board_VLSFO_In_MT",
    "Remaining_On_Board_LSMGO_In_MT",
    "Remaining_On_Board_ULSGO_In_MT",
    "Remaining_On_Board_VLSGO_In_MT",
)


def _noonreportdata_from_row(cell: Any) -> dict[str, Any]:
    """Pull EOV-relevant numeric fields; keys match voyagepm_be noonreportdata tags."""
    out: dict[str, Any] = {}
    for key in _EOV_ND_KEYS:
        # openpyxl headers are lower_snake; BE tags are mixed — try both
        raw = cell(key.lower(), key)
        if raw is None:
            continue
        num = _float_or_none(raw)
        out[key] = num if num is not None else raw
    # Prefer Avg_Speed from LOG_SPEED / Average_Speed_Since_SOV if missing
    if out.get("Avg_Speed") is None:
        alt = _float_or_none(cell("avg_speed", "average_speed_since_sov", "log_speed"))
        if alt is not None:
            out["Avg_Speed"] = alt
    if out.get("Wind_Force") is None:
        bf = _float_or_none(cell("beaufort_scale", "wind_force"))
        if bf is not None:
            out["Wind_Force"] = bf
    return out


def is_arrival_report(report_type: str | None) -> bool:
    t = (report_type or "").strip().lower()
    return "arrival" in t


def is_departure_report(report_type: str | None) -> bool:
    t = (report_type or "").strip().lower()
    return "departure" in t and "pre" not in t


def voyage_has_departed(voyage_rec: dict[str, Any] | None) -> bool:
    """True once a real Departure Report has been ingested (not pre-voyage)."""
    rec = voyage_rec or {}
    if rec.get("passage_weather_active"):
        return True
    if is_departure_report((rec.get("last_noon") or {}).get("report_type")):
        return True
    for row in rec.get("noon_history") or []:
        if is_departure_report(row.get("report_type")):
            return True
    return False


if __name__ == "__main__":
    rows = [
        {"noon_id": "b", "observed_at": "2026-01-02T00:00:00Z"},
        {"noon_id": "a", "observed_at": "2026-01-01T00:00:00Z"},
        {"noon_id": "c", "observed_at": None},
    ]
    assert [r["noon_id"] for r in sort_noon_rows(rows)] == ["a", "b", "c"]
    assert is_departure_report("Departure Report")
    assert not is_departure_report("Noon Report")
    assert not is_departure_report("pre-voyage departure")
    assert voyage_has_departed({"noon_history": [{"report_type": "Departure Report"}]})
    assert not voyage_has_departed({"last_noon": {"report_type": "Noon Report"}})
    print("noon_io self-check ok")
