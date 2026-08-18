"""Inbox drop-folder I/O — pre-voyage / noon Excel (or CSV stand-in).

Supports:
  1) Flat CSV/xlsx demo schema (voyage_number, cp_speed_kn, waypoints, …)
  2) Real SSH Pre-Dep Voyage workbook (sheets: Vessel Details, Voyage Details,
     Waypoints List, CP Terms FWC)

Drop pre-voyage files into VPM_INBOX_DIR/incoming/. Noon → VPM_NOON_INBOX_DIR/incoming/.
"""

from __future__ import annotations

import csv
import re
import shutil
from pathlib import Path
from typing import Any, Literal

from vpm_agents.tools.voyage_registry import normalize_voyage_number

from vpm_agents.tools.folder_layout import FAILED, INCOMING, SENT, ensure_drop_dirs, incoming_dir

MsgKind = Literal["pre_voyage", "noon_report", "unknown"]

_INBOX_SUFFIXES = {".csv", ".xlsx", ".xlsm"}

# Section tags — never treat as field labels (RPM tables use Ballast/Laden as row groups)
_SECTION_TAGS = {
    "ballast",
    "laden",
    "remarks",
    "engine info",
    "general vessel info",
    "charter party info. for voyage",
    "intial voyage info.",
    "initial voyage info.",
    "allowed weather",
    "sample format",
}

_LAT_HEADERS = {"lat", "latitude"}
_LON_HEADERS = {"long", "lon", "lng", "longitude"}


def _norm_header(h: str) -> str:
    return h.strip().lower().replace(" ", "_")


def _norm_label(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _is_xlsx(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm"}


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl required for .xlsx — pip install openpyxl") from e
    return load_workbook(path, data_only=True)


def _sheet_role(title: str) -> str | None:
    t = _norm_label(title)
    if "waypoint" in t:
        return "waypoints"
    if "voyage" in t:
        return "voyage"
    if "vessel" in t:
        return "vessel"
    if "cp term" in t or "fwc" in t or "charter party" in t:
        return "cp"
    return None


def _sheet_map(wb) -> dict[str, Any]:
    """Map logical roles → worksheet by title keywords, not a fixed ship template."""
    out: dict[str, Any] = {}
    for ws in wb.worksheets:
        role = _sheet_role(ws.title)
        if role and role not in out:
            out[role] = ws
    return out


def is_predep_workbook(path: Path) -> bool:
    if not _is_xlsx(path):
        return False
    wb = _load_wb(path)
    try:
        sheets = _sheet_map(wb)
        if "waypoints" in sheets:
            return True
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            return False
        # Combined noon Excel: Voyage_Number + Latitude columns — not a Pre-Dep form
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
        cols = {_norm_header(str(h)) for h in header if h is not None}
        if _kind_from_cols(cols) == "noon_report":
            return False
        return _find_lat_lon_cols(ws)[0] is not None
    finally:
        wb.close()


def _as_float(val: Any) -> float | None:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
        return float(s)
    return None


def _is_field_label(cell: Any) -> bool:
    if cell is None or not isinstance(cell, str):
        return False
    key = _norm_label(cell)
    if len(key) < 3 or key in _SECTION_TAGS:
        return False
    if key in _LAT_HEADERS or key in _LON_HEADERS:
        return False
    return bool(re.search(r"[a-zA-Z]", key))


def _sheet_fields(ws) -> dict[str, list[Any]]:
    """Label → values to the right. Works for col-C labels or any leading field name."""
    found: dict[str, list[Any]] = {}
    max_row = min(ws.max_row or 1, 80)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=12, values_only=True):
        cells = list(row)
        label_i = None
        for i, c in enumerate(cells[:3]):
            if _is_field_label(c):
                label_i = i
        if label_i is None:
            continue
        rest = [c for c in cells[label_i + 1 :] if c is not None and str(c).strip() != ""]
        if not rest:
            continue
        found.setdefault(_norm_label(cells[label_i]), rest)
    return found


def _merge_fields(wb, sheets: dict[str, Any]) -> dict[str, list[Any]]:
    fields: dict[str, list[Any]] = {}
    for role in ("voyage", "cp", "vessel"):
        ws = sheets.get(role)
        if ws is None:
            continue
        for k, v in _sheet_fields(ws).items():
            fields.setdefault(k, v)
    for ws in wb.worksheets:
        for k, v in _sheet_fields(ws).items():
            fields.setdefault(k, v)
    return fields


def _pick(mapping: dict[str, list[Any]], *needles: str, numeric: bool = False) -> Any:
    """Best label containing a needle; skip allowance/variance extras unless asked."""
    ranked: list[tuple[int, int, str, list[Any]]] = []
    for needle in needles:
        n = _norm_label(needle)
        for k, vals in mapping.items():
            if n not in k:
                continue
            extra = k.replace(n, "", 1).strip()
            if extra and any(w in extra for w in ("allowance", "variance", "allowed")):
                continue
            ranked.append((len(extra), -len(n), k, vals))
    if not ranked:
        return None
    ranked.sort()
    vals = ranked[0][3]
    if numeric:
        return _as_float(vals[0])
    return vals[0]


def _find_lat_lon_cols(ws) -> tuple[int | None, int | None, int, int]:
    """Return (lat_col, lon_col, name_col, first_data_row)."""
    if ws is None:
        return None, None, 0, 1
    lat_col = lon_col = None
    name_col = 0
    header_row = 1
    for r_i, row in enumerate(ws.iter_rows(max_row=min(ws.max_row or 1, 25), max_col=12, values_only=True), 1):
        for ci, c in enumerate(row):
            if c is None:
                continue
            n = _norm_label(c).rstrip(".")
            if n in _LAT_HEADERS:
                lat_col = ci
                header_row = r_i
            elif n in _LON_HEADERS:
                lon_col = ci
                header_row = r_i
            elif n in {"wpt no", "wpt", "waypoint", "name"}:
                name_col = ci
        if lat_col is not None and lon_col is not None:
            return lat_col, lon_col, name_col, header_row + 1
    return None, None, 0, 1


def _parse_waypoints(ws) -> tuple[list[list[float]], list[str]]:
    lat_col, lon_col, name_col, start = _find_lat_lon_cols(ws)
    if lat_col is None or lon_col is None:
        return [], []
    waypoints: list[list[float]] = []
    wp_names: list[str] = []
    need = max(lat_col, lon_col, name_col) + 1
    for row in ws.iter_rows(min_row=start, max_row=ws.max_row or start, max_col=need, values_only=True):
        cells = list(row) + [None] * need
        lat_s, lon_s = cells[lat_col], cells[lon_col]
        if lat_s is None or lon_s is None:
            continue
        if str(lat_s).strip() == "" or str(lon_s).strip() == "":
            continue
        if _norm_label(str(lat_s)) in _LAT_HEADERS:
            continue
        try:
            lat = parse_dm_coordinate(lat_s)
            lon = parse_dm_coordinate(lon_s)
        except ValueError:
            continue
        waypoints.append([lat, lon])
        name = cells[name_col]
        wp_names.append("" if name is None else str(name).strip())
    return waypoints, wp_names


def parse_dm_coordinate(text: str) -> float:
    """Parse '12 39.812 N' / '109 24.414 E' (deg + decimal minutes + hemisphere)."""
    s = str(text).strip().upper().replace("°", " ").replace("'", " ").replace('"', " ")
    s = re.sub(r"\s+", " ", s)
    m = re.match(
        r"^(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s*([NSEW])$",
        s,
    )
    if not m:
        # already decimal?
        try:
            return float(s)
        except ValueError as e:
            raise ValueError(f"bad coordinate: {text!r}") from e
    deg = float(m.group(1))
    minutes = float(m.group(2))
    hemi = m.group(3)
    val = deg + minutes / 60.0
    if hemi in ("S", "W"):
        val = -val
    return round(val, 6)


def parse_predep_workbook(path: Path) -> dict[str, Any]:
    """Parse a Pre-Dep-style workbook: any sheet titles, label/value rows, lat/long table."""
    wb = _load_wb(path)
    try:
        sheets = _sheet_map(wb)
        fields = _merge_fields(wb, sheets)
        wp_ws = sheets.get("waypoints") or next(
            (ws for ws in wb.worksheets if _find_lat_lon_cols(ws)[0] is not None),
            None,
        )
        if wp_ws is None:
            raise ValueError(f"{path.name}: no waypoint sheet with Lat/Long columns")

        voyage_number = _pick(fields, "voyage number", "voyage no")
        source_port = _pick(fields, "departure port", "from port", "origin")
        dest_port = _pick(fields, "destination port", "dest port", "arrival port")
        vessel_name = _pick(fields, "vessel name", "ship name")
        imo = _pick(fields, "imo no", "imo number")
        cp_speed = _pick(fields, "cp speed in kts", "cp speed", "speed in kts", numeric=True)
        cp_consumption = _pick(
            fields, "cp consumption", "consumptions(total)", "consumption mt/day", numeric=True
        )

        if voyage_number is None:
            raise ValueError(f"{path.name}: Voyage Number missing")
        if source_port is None or dest_port is None:
            raise ValueError(f"{path.name}: Departure/Destination port missing")
        if cp_speed is None:
            raise ValueError(f"{path.name}: CP Speed missing (need a numeric knots value)")

        waypoints, wp_names = _parse_waypoints(wp_ws)
        if len(waypoints) < 2:
            raise ValueError(f"{path.name}: waypoint list needs ≥2 Lat/Long points")

        vessel_id = str(imo).strip() if imo is not None else (str(vessel_name).strip() if vessel_name else "")
        cond = _pick(fields, "condition (ballast/laden)", "condition")
        return {
            "voyage_number": normalize_voyage_number(str(voyage_number)),
            "vessel_id": vessel_id,
            "vessel_name": "" if vessel_name is None else str(vessel_name).strip(),
            "source_port": str(source_port).strip(),
            "dest_port": str(dest_port).strip(),
            "cp_speed_kn": float(cp_speed),
            "cp_consumption_mt_day": cp_consumption,
            "alert_emails": [],
            "master_waypoints": waypoints,
            "waypoint_names": wp_names,
            "source_file": str(path),
            "format": "predep_xlsx",
            "condition": "" if cond is None else str(cond).strip(),
            "etd": str(_pick(fields, "estimated departure time") or "").strip(),
            "eta": str(_pick(fields, "estimated arrival time") or "").strip(),
            "displacement": _pick(fields, "displacement", numeric=True),
            "cargo_weight": _pick(fields, "cargo weight", numeric=True),
            "max_draft_on_departure": _pick(
                fields, "max draft upon departure", "max draft on departure", numeric=True
            ),
            "voyage_priority": str(_pick(fields, "voyage priority") or "").strip(),
        }
    finally:
        wb.close()


def _read_table(path: Path) -> tuple[list[str], list[str]]:
    """Return (headers, first_data_row_values) from flat csv or single-sheet xlsx."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if len(rows) < 2:
            raise ValueError(f"{path.name}: need header + one data row")
        return [_norm_header(h) for h in rows[0]], [c.strip() for c in rows[1]]

    if suffix in {".xlsx", ".xlsm"}:
        wb = _load_wb(path)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            data_row = next(rows_iter, None)
            if not header_row or not data_row:
                raise ValueError(f"{path.name}: need header + one data row")
            headers = [_norm_header(str(h)) for h in header_row if h is not None]
            values = ["" if v is None else str(v).strip() for v in data_row[: len(headers)]]
            return headers, values
        finally:
            wb.close()

    raise ValueError(f"unsupported inbox file type: {suffix}")


def _kind_from_cols(cols: set[str]) -> MsgKind | None:
    if "waypoints" in cols and "cp_speed_kn" in cols:
        return "pre_voyage"
    has_lat = bool(cols & {"lat", "latitude"})
    has_lon = bool(cols & {"lon", "long", "lng", "longitude"})
    if has_lat and has_lon and "voyage_number" in cols:
        return "noon_report"
    return None


def classify_inbox_file(path: str | Path) -> MsgKind:
    path = Path(path)
    name_l = path.name.lower()

    # Multi-sheet Pre-Dep first (waypoint sheet). Noon workbooks also have Lat/Long
    # but usually no "waypoints" sheet — those must classify as noon_report.
    if _is_xlsx(path):
        try:
            if is_predep_workbook(path):
                return "pre_voyage"
        except Exception:
            pass

    try:
        headers, _ = _read_table(path)
    except Exception:
        headers = []
    kind = _kind_from_cols(set(headers))
    if kind:
        return kind

    if _is_xlsx(path) and (
        name_l.startswith("pre-dep") or "pre-dep voyage" in name_l or "pre_dep" in name_l
    ):
        return "pre_voyage"  # let parse raise a clear error
    return "unknown"


def _parse_waypoint_field(wp_raw: str) -> list[list[float]]:
    waypoints: list[list[float]] = []
    for pair in wp_raw.split(";"):
        pair = pair.strip()
        if not pair:
            continue
        lat_s, lon_s = pair.split(",")
        waypoints.append([float(lat_s), float(lon_s)])
    if len(waypoints) < 2:
        raise ValueError("pre-voyage needs ≥2 waypoints")
    return waypoints


def parse_pre_voyage(path: str | Path) -> dict[str, Any]:
    """Parse flat CSV/xlsx demo schema OR real Pre-Dep Voyage workbook."""
    path = Path(path)

    if _is_xlsx(path) and is_predep_workbook(path):
        return parse_predep_workbook(path)

    if path.suffix.lower() == ".csv":
        lines = path.read_text(encoding="utf-8-sig").strip().splitlines()
        if len(lines) < 2:
            raise ValueError(f"{path.name}: need header + one data row")
        headers = [_norm_header(h) for h in lines[0].split(",")]
        parts = lines[1].split(",", 6)
        if len(parts) < 7:
            raise ValueError(f"{path.name}: bad pre-voyage CSV row")
        row = dict(zip(headers[:6], [p.strip() for p in parts[:6]]))
        row["waypoints"] = parts[6].strip()
    else:
        headers, values = _read_table(path)
        row = dict(zip(headers, values))

    for key in ("voyage_number", "vessel_id", "source_port", "dest_port", "cp_speed_kn", "waypoints"):
        if key not in row or not row[key]:
            raise ValueError(f"pre-voyage missing {key}")
    emails = [e for e in row.get("alert_emails", "").split("|") if e.strip()]
    cons_raw = row.get("cp_consumption_mt_day") or row.get("cp_consumption")
    cons = None
    if cons_raw not in (None, ""):
        try:
            cons = float(cons_raw)
        except ValueError:
            cons = None
    return {
        "voyage_number": normalize_voyage_number(row["voyage_number"]),
        "vessel_id": row["vessel_id"],
        "source_port": row["source_port"],
        "dest_port": row["dest_port"],
        "cp_speed_kn": float(row["cp_speed_kn"]),
        "cp_consumption_mt_day": cons,
        "alert_emails": emails,
        "master_waypoints": _parse_waypoint_field(row["waypoints"]),
        "source_file": str(path),
        "format": "flat",
    }


def parse_noon_report(path: str | Path) -> dict[str, Any]:
    """Columns: voyage_number, lat, lon  (optional observed_at)."""
    headers, values = _read_table(Path(path))
    row = dict(zip(headers, values))
    for key in ("voyage_number", "lat", "lon"):
        if key not in row or not row[key]:
            raise ValueError(f"noon report missing {key}")
    return {
        "voyage_number": normalize_voyage_number(row["voyage_number"]),
        "lat": float(row["lat"]),
        "lon": float(row["lon"]),
        "observed_at": row.get("observed_at") or None,
        "source_file": str(path),
    }


def list_inbox(inbox_dir: str | Path) -> list[Path]:
    inbox = Path(inbox_dir)
    ensure_drop_dirs(inbox)
    drop = incoming_dir(inbox)
    return sorted(
        p
        for p in drop.iterdir()
        if p.is_file() and p.suffix.lower() in _INBOX_SUFFIXES and not p.name.startswith(".")
    )


def relocate_inbox_file(path: Path, dest_dir: str | Path) -> Path:
    path = Path(path)
    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / path.name
    if dest.exists():
        dest = dest_dir / f"{path.stem}_{path.stat().st_mtime_ns}{path.suffix}"
    shutil.move(str(path), str(dest))
    return dest


def archive_inbox_file(path: Path, dest_subdir: str = SENT) -> Path:
    return relocate_inbox_file(path, path.parent.parent / dest_subdir)


if __name__ == "__main__":
    assert _as_float("Ballast") is None
    assert _as_float(10.7) == 10.7
    assert _as_float("12.5") == 12.5
    mapping = {
        "cp speed in kts/day": [10.7, 12.5],
        "cp speed allowance variance (kts)": [0.5],
        "cp consumptions(total) in mts/day": ["Ballast", 25],
    }
    assert _pick(mapping, "cp speed", numeric=True) == 10.7
    assert _pick(mapping, "cp consumption", numeric=True) is None
    assert _kind_from_cols({"voyage_number", "latitude", "longitude"}) == "noon_report"
    assert _kind_from_cols({"voyage_number", "lat", "lon"}) == "noon_report"
    assert _kind_from_cols({"waypoints", "cp_speed_kn"}) == "pre_voyage"
    print("inbox_io self-check ok")
